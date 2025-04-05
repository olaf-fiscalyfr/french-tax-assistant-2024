import streamlit as st
import pandas as pd
import json
import time
import re
import os
from pathlib import Path
import io
from dotenv import load_dotenv
from openai import OpenAI
import tempfile

# For PDF extraction
from PyPDF2 import PdfReader
# For Word documents
from docx import Document
# For Excel files
import openpyxl
# For image processing
from PIL import Image

# Load environment variables from .env file if it exists
load_dotenv()

# Get API key from environment or use a placeholder
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")

# Constants
CURRENT_YEAR = 2024
PREVIOUS_YEAR = CURRENT_YEAR - 1
EXCHANGE_RATES = {
    "GBP": 1.1812,
    "USD": 0.9204,
    "CHF": 1.0418,
    "CAD": 0.6812,
    "EUR": 1.0000
}

# French tax form and code mapping
FORM_DESCRIPTIONS = {
    "2042": "Déclaration des revenus",
    "2044": "Déclaration des revenus fonciers",
    "2047": "Déclaration des revenus encaissés à l'étranger",
    "3916": "Déclaration des comptes ouverts à l'étranger",
    "2086": "Déclaration des frais professionnels"
}

# Define common tax codes and descriptions for reference and validation
TAX_CODES = {
    # Form 2042 - Common income declaration
    "2042": {
        "1AJ": "Salaries - Declarant 1",
        "1BJ": "Salaries - Declarant 2",
        "1AP": "Pensions, Retirement - Declarant 1",
        "1BP": "Pensions, Retirement - Declarant 2",
        "2TR": "Interest income subject to flat tax (PFU)",
        "2TS": "Interest from regulated savings accounts (Livret A, etc.)",
        "2BH": "French bank interest - Declarant 1",
        "2CH": "French bank interest - Declarant 2",
        "2DC": "Dividends subject to progressive tax",
        "2EE": "Capital gains on securities",
        "5HQ": "Micro-entrepreneur/Auto-entrepreneur income",
        "5KO": "Miscellaneous income - Declarant 1",
        "5LO": "Miscellaneous income - Declarant 2",
        "5TE": "Micro-BIC income (Chambre d'hôte/gites)",
        "7DB": "Energy production credits (solar panels, etc.)",
        "7DQ": "Energy equipment installations (tax credit)"
    },
    # Form 2044 - Real estate income
    "2044": {
        "4BA": "Gross rental income",
        "4BB": "Property tax paid",
        "4BC": "Deductible interest on loans",
        "4BH": "Property expenses",
        "4BK": "Property management fees",
        "4BL": "Insurance premiums"
    },
    # Form 2047 - Foreign income
    "2047": {
        "1AF": "Foreign pensions - Declarant 1",
        "1BF": "Foreign pensions - Declarant 2",
        "1AG": "Government/Civil Service pensions - Declarant 1",
        "1BG": "Government/Civil Service pensions - Declarant 2",
        "2AB": "Foreign dividends",
        "2BG": "Foreign interest income - Declarant 1",
        "2CG": "Foreign interest income - Declarant 2",
        "3VG": "Foreign capital gains",
        "3VH": "Foreign capital gains exempt but used for rate calculation"
    },
    # Form 3916 - Foreign accounts
    "3916": {
        "8UU": "Foreign bank account - Country code",
        "8TK": "Foreign bank account - Account number",
        "8QS": "Foreign bank account - Name of institution",
        "8RT": "Foreign bank account - Full address",
        "8QU": "Foreign bank account - Account type (current/savings)"
    },
    # Form 2086 - Professional expenses
    "2086": {
        "AK": "Professional expenses - Declarant 1",
        "BK": "Professional expenses - Declarant 2",
        "WW": "Home-to-work distance (km)"
    }
}

def extract_text_from_file(file):
    """Extract text from PDF, DOCX, TXT or other supported files"""
    suffix = Path(file.name).suffix.lower()
    
    if suffix == ".pdf":
        try:
            # Extract text directly from PDF
            reader = PdfReader(file)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            return text
        except Exception as e:
            st.error(f"Error extracting text from PDF: {e}")
            return ""
            
    elif suffix == ".docx":
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as tmp:
                tmp.write(file.getvalue())
                tmp_path = tmp.name
            
            doc = Document(tmp_path)
            text = "\n".join(p.text for p in doc.paragraphs)
            
            # Clean up the temp file
            os.unlink(tmp_path)
            return text
        except Exception as e:
            st.error(f"Error extracting text from DOCX: {e}")
            return ""
    
    elif suffix == ".xlsx" or suffix == ".xls":
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(file.getvalue())
                tmp_path = tmp.name
            
            # Load workbook
            wb = openpyxl.load_workbook(tmp_path)
            
            # Extract text from each sheet
            texts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_text = f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        sheet_text += row_text + "\n"
                
                texts.append(sheet_text)
            
            # Clean up the temp file
            os.unlink(tmp_path)
            return "\n\n".join(texts)
        except Exception as e:
            st.error(f"Error extracting text from Excel file: {e}")
            return ""
            
    elif suffix == ".txt":
        return file.getvalue().decode("utf-8", errors="ignore")
    
    elif suffix == ".json":
        try:
            json_data = json.loads(file.getvalue())
            # Format the JSON as a string
            return json.dumps(json_data, indent=2, ensure_ascii=False)
        except Exception as e:
            st.error(f"Error parsing JSON file: {e}")
            return file.getvalue().decode("utf-8", errors="ignore")
        
    else:
        st.warning(f"Unsupported file type: {suffix}")
        return ""

def extract_with_gpt(text, additional_context=""):
    """Extract tax information using GPT-4"""
    if not OPENAI_API_KEY:
        st.error("OpenAI API Key is required. Please enter it in the sidebar.")
        return None
        
    client = OpenAI(api_key=OPENAI_API_KEY)
    
    exchange_rates_text = "\n".join([f"- 1 {currency} = {rate} EUR" for currency, rate in EXCHANGE_RATES.items()])
    
    # Enhanced prompt with more comprehensive tax information
    prompt = f"""
You are a professional French tax assistant specializing in expatriate and international taxation for {CURRENT_YEAR}.

Extract ALL tax-relevant income data from the following document. Be thorough and precise.

Return your response ONLY as a JSON list of income entries with the following structure:
[
    {{
        "Description": "UK Civil Service Pension",
        "Form": "2047",
        "Code": "1AG",
        "Source Currency": "GBP",
        "Amount (Source)": 9351,
        "Amount (€)": 11045.52,
        "Taxable in France (€)": "",
        "Taxable in Source Country (€)": 11045.52,
        "Notes": "Exempt in France, taxable in UK per tax treaty"
    }},
    {{
        "Description": "Bank Interest",
        "Form": "2042",
        "Code": "2TR",
        "Source Currency": "EUR",
        "Amount (Source)": 350,
        "Amount (€)": 350,
        "Taxable in France (€)": 350,
        "Taxable in Source Country (€)": "",
        "Notes": "Subject to PFU flat tax at 30%"
    }},
    {{
        "Description": "Chambre d'Hôte Income",
        "Form": "2042",
        "Code": "5TE", 
        "Source Currency": "EUR",
        "Amount (Source)": 7387,
        "Amount (€)": 7387,
        "Taxable in France (€)": 7387,
        "Taxable in Source Country (€)": "",
        "Notes": "Micro-BIC regime, 71% abatement applies"
    }}
]

Include ALL of the following types of income that appear in the document:
1. Employment income/salaries (Form 2042, codes 1AJ, 1BJ, etc.)
2. Pensions and retirement income (Form 2042, codes 1AP, 1BP, etc. or Form 2047 for foreign)
   a. Use code 1AG/1BG for government/civil service pensions (UK Civil Service, etc.)
   b. Use code 1AF/1BF for standard foreign pensions
3. Investment income - dividends, interest (Form 2042, codes 2DC, 2TR, etc.)
4. Real estate income (Form 2044, codes 4BA, etc.)
5. Capital gains (Form 2042, codes 3VG, 2EE, etc.)
6. Foreign income of all types (Form 2047)
7. Foreign bank accounts (Form 3916)
8. Micro-entrepreneur or chambres d'hôte income (Form 2042, codes 5HQ, 5TE)
9. Income from energy production like solar panels (Form 2042, code 7DB)

Important tax form codes:
- 2042: Standard French income declaration
- 2044: Real estate income (unfurnished rentals)
- 2047: Foreign income declaration
- 3916: Foreign bank accounts declaration

Use these exchange rates:
{exchange_rates_text}

Additional context about this taxpayer:
{additional_context}

IMPORTANT: Return ONLY the JSON array. No explanations, no markdown formatting.

Document content:
{text}
"""
    
    try:
        response = client.chat.completions.create(
            model="gpt-4-turbo",
            messages=[
                {"role": "system", "content": "You are a professional French tax assistant that outputs precise, structured data."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=4000,
            temperature=0.2  # Lower temperature for more consistent outputs
        )
        return response.choices[0].message.content
    except Exception as e:
        st.error(f"OpenAI API Error: {e}")
        return None

def normalize_code(code):
    """Normalize tax code to uppercase with no spaces or special characters"""
    return re.sub(r"[^0-9A-Z]", "", code.upper())

def safe_parse_json(gpt_output):
    """Safely parse GPT output to JSON, with fallbacks for different output formats"""
    try:
        # Try direct parsing
        if gpt_output and gpt_output.strip():
            # Clean the output to extract just the JSON array
            clean_output = gpt_output.strip()
            
            # If output is wrapped in code block, extract just the JSON 
            if "```json" in clean_output:
                match = re.search(r"```json\s*(\[.*?\])\s*```", clean_output, re.DOTALL)
                if match:
                    clean_output = match.group(1)
            # If output is not a direct array, search for array pattern
            elif not clean_output.startswith("["):
                match = re.search(r"\[\s*{.*?}\s*(?:,\s*{.*?}\s*)*\]", clean_output, re.DOTALL)
                if match:
                    clean_output = match.group(0)
                    
            return json.loads(clean_output)
        return []
    except Exception as e:
        st.error(f"JSON Parse Error: {e}")
        st.error("Raw output from GPT could not be parsed as JSON. Please try again.")
        return None

def validate_and_clean_data(parsed_data):
    """Validate and clean the parsed data"""
    if not parsed_data:
        return [], []
        
    valid_entries = []
    warning_messages = []
    
    for entry in parsed_data:
        # Ensure all required fields exist
        required_fields = ["Description", "Form", "Code", "Source Currency", "Amount (Source)"]
        
        if not all(field in entry for field in required_fields):
            missing_fields = [field for field in required_fields if field not in entry]
            warning_messages.append(f"Entry '{entry.get('Description', 'Unknown')}' is missing required fields: {', '.join(missing_fields)}")
            continue
            
        # Normalize and validate form and code
        entry["Form"] = entry["Form"].strip()
        entry["Code"] = normalize_code(entry["Code"])
        
        # Handle currencies and amounts
        source_currency = entry["Source Currency"].upper()
        source_amount = entry.get("Amount (Source)", 0)
        
        # Convert to float if string
        if isinstance(source_amount, str):
            try:
                source_amount = float(source_amount.replace(",", ".").replace(" ", ""))
            except ValueError:
                warning_messages.append(f"Could not convert amount '{source_amount}' for '{entry.get('Description', 'Unknown')}' to a number")
                source_amount = 0
            
        entry["Amount (Source)"] = source_amount
        
        # Calculate EUR amount if needed
        if source_currency != "EUR":
            exchange_rate = EXCHANGE_RATES.get(source_currency)
            if exchange_rate:
                entry["Amount (€)"] = round(source_amount * exchange_rate, 2)
            else:
                warning_messages.append(f"Unknown currency '{source_currency}' for '{entry.get('Description', 'Unknown')}'. Using 1:1 exchange rate.")
                entry["Amount (€)"] = source_amount
        else:
            entry["Amount (€)"] = source_amount
            
        # Ensure taxable fields exist and are numeric
        for field in ["Taxable in France (€)", "Taxable in Source Country (€)"]:
            if field not in entry or entry[field] == "":
                entry[field] = ""
            elif isinstance(entry[field], str):
                try:
                    entry[field] = float(entry[field].replace(",", ".").replace(" ", ""))
                except ValueError:
                    warning_messages.append(f"Could not convert {field} '{entry[field]}' for '{entry.get('Description', 'Unknown')}' to a number")
                    entry[field] = ""
        
        # Add if code exists in our reference (basic validation)
        if entry["Form"] in TAX_CODES and entry["Code"] in TAX_CODES[entry["Form"]]:
            entry["Valid Code"] = True
            entry["Code Description"] = TAX_CODES[entry["Form"]][entry["Code"]]
        else:
            entry["Valid Code"] = False
            if entry["Form"] in TAX_CODES:
                warning_messages.append(f"Unknown code '{entry['Code']}' for form '{entry['Form']}' - '{entry.get('Description', 'Unknown')}'")
            else:
                warning_messages.append(f"Unknown form '{entry['Form']}' for '{entry.get('Description', 'Unknown')}'")
            entry["Code Description"] = ""
        
        # Special rules for specific forms and codes
        if entry["Form"] == "2042" and entry["Code"] == "5TE":
            # Apply 71% automatic abatement for Chambre d'hôte income
            if "Taxable in France (€)" in entry and entry["Taxable in France (€)"]:
                raw_amount = float(entry["Taxable in France (€)"])
                entry["Abatement Amount"] = round(raw_amount * 0.71, 2)
                entry["Taxable in France (€) After Abatement"] = round(raw_amount * 0.29, 2)
                entry["Notes"] = entry.get("Notes", "") + " Micro-BIC regime, 71% abatement applies."
        
        # Add the entry to our validated list
        valid_entries.append(entry)
    
    return valid_entries, warning_messages

def create_excel_output(df, taxpayer_info=None, tax_year=CURRENT_YEAR):
    """Create a formatted Excel file with the tax data"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Main data sheet
        df.to_excel(writer, sheet_name=f"IRPP {tax_year}", index=False)
        
        # Format workbook (basic implementation)
        workbook = writer.book
        worksheet = writer.sheets[f"IRPP {tax_year}"]
        
        # Set column widths
        worksheet.set_column('A:A', 30)  # Description
        worksheet.set_column('B:B', 8)   # Form
        worksheet.set_column('C:C', 8)   # Code
        worksheet.set_column('D:D', 25)  # Code Description
        worksheet.set_column('E:E', 10)  # Source Currency
        worksheet.set_column('F:F', 15)  # Amount (Source)
        worksheet.set_column('G:G', 15)  # Amount (€)
        worksheet.set_column('H:H', 15)  # Taxable in France (€)
        worksheet.set_column('I:I', 15)  # Taxable in Source Country (€)
        worksheet.set_column('J:J', 40)  # Notes
    
    # Return the Excel file as bytes
    output.seek(0)
    return output.getvalue()

def create_json_output(df):
    """Create JSON output file for Clickimpôts"""
    # Only include rows that have values
    if len(df) == 0:
        return "[]"
        
    filtered_df = df[df['Taxable in France (€)'].astype(str).str.strip() != '' | 
                     df['Taxable in Source Country (€)'].astype(str).str.strip() != '']
                     
    json_data = []
    
    for _, row in filtered_df.iterrows():
        value = row['Taxable in France (€)'] or row['Taxable in Source Country (€)']
        
        # Skip if no value
        if pd.isna(value) or value == "":
            continue
            
        # Ensure value is numeric
        if isinstance(value, str):
            try:
                value = float(value.replace(',', '.').replace(' ', ''))
            except ValueError:
                continue
                
        json_data.append({
            "form": row['Form'],
            "code": row['Code'],
            "description": row['Description'],
            "value": float(value) if isinstance(value, (int, float)) else 0
        })
        
    return json.dumps(json_data, indent=2, ensure_ascii=False)

def extract_bank_accounts(text):
    """Extract bank account information from the document"""
    bank_accounts = []
    
    # Look for table format
    table_pattern = r"(?i)(name\s+of\s+institution|institution|bank).*?(account\s+number|account|number).*?(type|date\s+opened)"
    
    if re.search(table_pattern, text, re.MULTILINE):
        lines = text.split('\n')
        in_table = False
        header_line = None
        
        for line in lines:
            if re.search(table_pattern, line, re.IGNORECASE):
                in_table = True
                header_line = line
                continue
                
            if in_table and line.strip() and len(line.split()) >= 2:
                # Try to extract the fields based on the header
                if header_line:
                    # Try to determine column positions
                    header_parts = re.split(r'\s{2,}|\t', header_line)
                    line_parts = re.split(r'\s{2,}|\t', line)
                    
                    if len(line_parts) >= 2:
                        # Basic extraction
                        institution = line_parts[0] if len(line_parts) > 0 else "Unknown"
                        account_number = line_parts[1] if len(line_parts) > 1 else "Unknown"
                        account_type = line_parts[2] if len(line_parts) > 2 else "Unknown"
                        
                        # Skip if it looks like a header or continuation
                        if institution.lower() in ["name", "institution", "bank", "n/a", "active accounts"]:
                            continue
                        
                        bank_accounts.append({
                            "institution": institution.strip(),
                            "account_number": account_number.strip(),
                            "account_type": account_type.strip()
                        })
            
            # End of table detection
            if in_table and not line.strip():
                in_table = False
    
    # If no table found, try regex patterns
    if not bank_accounts:
        # Match individual bank account entries
        account_patterns = [
            r"(?i)(?:bank|institution)[\s:]+([^\n,]+)[\s,]*(?:account|number)[\s:]+([^\n,]+)",
            r"(?i)(Credit Agricole|HSBC|Santander|Barclays|Revolut)[\s\w]*?[\s:]+([A-Z0-9]{8,})"
        ]
        
        for pattern in account_patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                bank_accounts.append({
                    "institution": match[0].strip(),
                    "account_number": match[1].strip(),
                    "account_type": "Unknown"
                })
    
    return bank_accounts

def main():
    
    st.set_page_config(
        page_title=f"🇫🇷 French Tax Assistant {CURRENT_YEAR}",
        page_icon="🇫🇷",
        layout="wide"
    )
    
    st.title(f"🇫🇷 French Tax Assistant {CURRENT_YEAR}")
    st.markdown("""
    Upload client tax documents to automatically extract tax information and generate French tax declaration files.
    """)
    
    # Initialize session state if not already done
    if 'extracted_data' not in st.session_state:
        st.session_state.extracted_data = None
    if 'uploaded_files' not in st.session_state:
        st.session_state.uploaded_files = []
    if 'all_text' not in st.session_state:
        st.session_state.all_text = ""
    if 'taxpayer_info' not in st.session_state:
        st.session_state.taxpayer_info = None
    if 'warnings' not in st.session_state:
        st.session_state.warnings = []
    
    # Sidebar
    with st.sidebar:
        st.header("Settings")
        api_key = st.text_input("OpenAI API Key", 
                               value="" if not OPENAI_API_KEY or OPENAI_API_KEY.startswith("sk-") else OPENAI_API_KEY, 
                               type="password")
        if api_key:
            OPENAI_API_KEY = api_key
            
        st.divider()
        st.markdown("### Tax Year")
        tax_year = st.selectbox("Select Tax Year", [CURRENT_YEAR, CURRENT_YEAR-1, CURRENT_YEAR-2], index=0)
        
        st.divider()
        st.markdown("### Exchange Rates")
        with st.expander("Edit Exchange Rates"):
            for currency, rate in EXCHANGE_RATES.items():
                if currency != "EUR":  # Skip EUR as it's always 1:1
                    new_rate = st.number_input(f"1 {currency} = € ", value=rate, step=0.0001, format="%.4f")
                    EXCHANGE_RATES[currency] = new_rate
            
        st.divider()
        st.markdown("### About")
        st.markdown("""
        This tool extracts tax data from documents using AI and prepares files for French tax declarations.
        
        Supported file types:
        - PDF
        - Word (.docx)
        - Excel (.xlsx, .xls)
        - Text (.txt)
        """)
        
        # Export Settings
        st.divider()
        st.markdown("### Export Options")
        export_format = st.radio(
            "Default Export Format",
            ["Excel", "JSON", "Both"],
            index=0
        )
    
    # Main content area
    col1, col2 = st.columns([1, 1])
    
    with col1:
        uploaded_files = st.file_uploader(
            "📁 Upload tax documents", 
            type=["pdf", "docx", "txt", "xlsx", "xls"], 
            accept_multiple_files=True,
            help="Upload client questionnaires and supporting documents",
            key="file_uploader"
        )
        
        if uploaded_files and uploaded_files != st.session_state.uploaded_files:
            st.session_state.uploaded_files = uploaded_files
            st.session_state.all_text = ""  # Reset text when new files are uploaded
    
    with col2:
        # Taxpayer information section
        st.subheader("Client Information")
        taxpayer_col1, taxpayer_col2 = st.columns(2)
        
        with taxpayer_col1:
            taxpayer_name = st.text_input("Client Name")
            taxpayer_address = st.text_input("Client Address")
        
        with taxpayer_col2:
            taxpayer_email = st.text_input("Client Email")
            taxpayer_phone = st.text_input("Client Phone")
        
        # Store taxpayer info in session state
        if taxpayer_name:
            st.session_state.taxpayer_info = {
                'name': taxpayer_name,
                'address': taxpayer_address,
                'email': taxpayer_email,
                'phone': taxpayer_phone,
                'tax_year': tax_year
            }
        
        # Additional context for analysis
        st.subheader("Additional Context")
        additional_context = st.text_area(
            "Add any additional information that might be helpful for the analysis",
            placeholder="Example: 'The client is a UK pensioner living in France since 2018, with Civil Service pension...'",
            help="This information will be used to improve the extraction accuracy"
        )
    
    # Process files when uploaded
    if st.session_state.uploaded_files:
        st.markdown("### 📄 Uploaded Documents")
        
        # Display the list of uploaded files
        for file in st.session_state.uploaded_files:
            file_size = len(file.getvalue()) / 1024  # Size in KB
            st.write(f"- {file.name} ({file_size:.1f} KB)")
        
        # Button to process documents
        if st.button("🔍 Analyze Documents", type="primary"):
            # Extract text from all documents
            all_text = ""
            progress_container = st.empty()
            progress_bar = progress_container.progress(0)
            
            for i, file in enumerate(st.session_state.uploaded_files):
                progress_text = st.empty()
                progress_text.text(f"Processing {file.name}...")
                
                # Extract text
                extracted_text = extract_text_from_file(file)
                
                all_text += "\n\n--- START OF DOCUMENT: " + file.name + " ---\n\n"
                all_text += extracted_text
                all_text += "\n\n--- END OF DOCUMENT: " + file.name + " ---\n\n"
                
                # Update progress
                progress = int((i + 1) / len(st.session_state.uploaded_files) * 50)
                progress_bar.progress(progress)
            
            st.session_state.all_text = all_text
            
            # Process with GPT
            progress_text.text("🧠 Analyzing documents with GPT-4 Turbo...")
            
            # Get additional context from taxpayer info if available
            context = additional_context
            if st.session_state.taxpayer_info:
                taxpayer_context = f"Client: {st.session_state.taxpayer_info.get('name', 'Unknown')}\n"
                taxpayer_context += f"Tax Year: {tax_year}\n"
                
                if context:
                    context = taxpayer_context + "\n" + context
                else:
                    context = taxpayer_context
            
            gpt_json = extract_with_gpt(all_text, context)
            progress_bar.progress(75)
            
            if gpt_json:
                progress_text.text("Parsing results...")
                parsed = safe_parse_json(gpt_json)
                
                if parsed:
                    progress_text.text("Validating and cleaning data...")
                    validated_data, warnings = validate_and_clean_data(parsed)
                    
                    # Store warnings in session state
                    st.session_state.warnings = warnings
                    
                    # Convert to DataFrame
                    df = pd.DataFrame(validated_data)
                    
                    # Ensure all required columns exist
                    required_columns = [
                        "Description", "Form", "Code", "Source Currency", 
                        "Amount (Source)", "Amount (€)", 
                        "Taxable in France (€)", "Taxable in Source Country (€)",
                        "Notes"
                    ]
                    
                    for col in required_columns:
                        if col not in df.columns:
                            df[col] = ""
                    
                    # Add code description column if missing
                    if "Code Description" not in df.columns:
                        df["Code Description"] = ""
                        # Populate code descriptions
                        for i, row in df.iterrows():
                            form = row["Form"]
                            code = row["Code"]
                            if form in TAX_CODES and code in TAX_CODES[form]:
                                df.at[i, "Code Description"] = TAX_CODES[form][code]


import streamlit as st

st.title("French Tax Assistant 2024")
st.write("Upload your tax documents below to extract relevant information.")

if 'uploaded_file' not in locals() and 'uploaded_file' not in globals():
    st.info("👈 Use the sidebar or uploader to begin.")